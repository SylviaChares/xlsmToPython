from openpyxl import Workbook
import pandas as pd

# 1. Tafeln einlesen (aus alter Datei)
df_tafeln = pd.read_excel('Tarifrechner_KLV_neu.xlsb', 
                          sheet_name='Tafeln', header=2)

# 2. Neue Arbeitsmappe erstellen
wb = Workbook()
wb.remove(wb.active)  # Leeres Blatt entfernen

# 3. Tafeln einfügen
ws_tafeln = wb.create_sheet('Tafeln')
for r_idx, row in enumerate(df_tafeln.values, start=4):
    for c_idx, value in enumerate(row, start=1):
        ws_tafeln.cell(row=r_idx, column=c_idx, value=value)

# 4. Kalkulation-Blatt erstellen
ws_kalk = wb.create_sheet('Kalkulation', 0)
# ... Formeln hier einfügen ...

# 5. Definierte Namen anlegen
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names['x'] = DefinedName('x', attr_text='Kalkulation!$B$4')
# ... weitere Namen ...

# 6. Speichern
wb.save('Tarifrechner_clean.xlsx')